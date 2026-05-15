"""
运行调度器 - 无人机配送调度系统
================================
支持命令行参数指定输入文件和输出目录。
生成飞行计划Excel和详细明细Excel。

用法:
    python run_scheduler.py                        # 使用默认路径 (input/input_data.xlsx, output/)
    python run_scheduler.py --input input/xxx.xlsx # 指定输入文件
    python run_scheduler.py -i xxx.xlsx -o ./out   # 简写
"""

import sys
import os
import argparse
import json
from typing import Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from drone_engine.map_data import get_map_data
from drone_engine.order import OrderManager
from drone_engine.scheduler import DroneScheduler
from drone_engine.optimizer import ALNSConfig
from drone_engine.flight_plan import FlightPlanExporter, FlightAction, ActionType
from drone_engine.drone import (
    DRONE_FIXED_COST,
    BATTERY_SWAP_COST,
    DRONE_COUNT,
    SIMULATION_DURATION_SECONDS,
)


def seconds_to_time_str(seconds: float) -> str:
    """秒数 -> '分:秒' 格式，秒数补零"""
    total_seconds = round(seconds)
    minutes = total_seconds // 60
    secs = total_seconds % 60
    return f"{minutes}:{secs:02d}"


def run(
    data_file: str = None,
    output_dir: str = None,
    iterations: int = 5000,
    time_limit: float = 120.0,
    max_orders: int = 350,
):
    """
    运行调度器。

    参数:
        data_file: 订单数据Excel文件路径（默认: input/input_data.xlsx）
        output_dir: 输出目录（默认: output/）
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    if data_file is None:
        input_dir = os.path.join(script_dir, "input")
        xlsx_files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx')]
        if not xlsx_files:
            raise FileNotFoundError(f"input/ 目录下没有 Excel 文件")
        data_file = os.path.join(input_dir, xlsx_files[0])
        print(f"  使用输入文件: {data_file}")
    if output_dir is None:
        output_dir = os.path.join(script_dir, "output")

    # 从文件名提取基础名称（不含扩展名）用于输出文件命名
    base_name = os.path.splitext(os.path.basename(data_file))[0]

    print("=" * 70)
    print(f"  无人机配送调度系统 - {base_name}")
    print("=" * 70)

    # 1. 加载订单数据
    print("\n[1/4] 加载订单数据...")
    order_manager = OrderManager()
    count = order_manager.load_from_excel(data_file, "订单任务")
    print(f"  加载订单数: {count}")
    print(f"  总潜在收入: {order_manager.get_total_potential_revenue():.2f} 元")

    # 类型分布
    stats = order_manager.get_type_statistics()
    print(f"  类型分布:")
    for otype, cnt in stats.items():
        print(f"    {otype.value}: {cnt}")

    # 2. 执行调度（ALNS优化）
    print("\n[2/4] 执行调度（ALNS优化）...")

    # ALNS参数配置
    alns_config = ALNSConfig(
        max_iterations=iterations,
        max_no_improve=500,
        time_limit_seconds=time_limit,
        destroy_min=5,
        destroy_max=30,
        initial_temperature=50.0,
        cooling_rate=0.997,
        battery_swap_threshold=20.0,
        max_orders_to_assign=max_orders,
        verbose=True,
        log_interval=100,
    )

    scheduler = DroneScheduler(
        map_data=get_map_data(),
        order_manager=order_manager,
        config=alns_config,
    )

    # 初始位置：3架无人机分别部署在3个陆地机场
    initial_positions = {
        "UAV-01": "测试基地L1",
        "UAV-02": "先导基地",
        "UAV-03": "秀水湾度假村",
    }

    result = scheduler.schedule(initial_positions=initial_positions)

    # 3. 输出结果摘要
    print("\n[3/4] 调度结果摘要:")
    print(f"  总订单数: {result['total_orders']}")
    print(f"  已配送订单数: {result['delivered_orders']}")
    print(f"  配送率: {result['delivered_orders'] / result['total_orders'] * 100:.1f}%")
    print(f"  总收入: {result['total_income']:.2f} 元")
    print(f"  固定成本: -{result['total_fixed_cost']:.2f} 元")
    print(
        f"  换电成本: -{result['total_swap_cost']:.2f} 元 ({result['total_swaps']}次)"
    )
    print(f"  未履约惩罚: -{result['unfulfilled_penalty']:.2f} 元")
    print(f"  总成本: {result['total_cost']:.2f} 元")
    print(f"  净利润: {result['net_profit']:.2f} 元")
    print(f"  总飞行距离: {result['total_distance_km']:.2f} km")

    # 各无人机统计
    print("\n  各无人机详情:")
    for drone_id, ds in result["drone_stats"].items():
        print(
            f"    {drone_id}: 配送{ds['total_deliveries']}单, "
            f"收入{ds['total_income']:.2f}元, "
            f"距离{ds['total_distance_km']:.2f}km, "
            f"换电{ds['total_swaps']}次, "
            f"最终位置={ds['final_location']}, "
            f"最终电量={ds['final_battery']:.1f}%"
        )

    # 4. 生成Excel文件
    print("\n[4/4] 生成Excel文件...")

    os.makedirs(output_dir, exist_ok=True)

    # 4a. 飞行计划Excel（3合1版）
    plan_file = os.path.join(output_dir, f"飞行计划_{base_name}.xlsx")
    export_flight_plan(result, plan_file)
    print(f"  飞行计划(合并): {plan_file}")

    # 4b. 3个独立无人机飞行计划Excel
    export_drone_flight_plans(result, output_dir, base_name)

    # 4c. 明细Excel
    detail_file = os.path.join(output_dir, f"调度明细_{base_name}.xlsx")
    export_details(result, detail_file, order_manager)
    print(f"  调度明细: {detail_file}")

    # 4d. 已分配订单列表
    order_list_file = os.path.join(output_dir, f"已分配订单_{base_name}.txt")
    export_assigned_orders(result, order_manager, order_list_file)
    print(f"  已分配订单: {order_list_file}")

    # 5. 生成JSON摘要（供前端使用）
    summary = {
        "status": "success",
        "input_file": data_file,
        "base_name": base_name,
        "total_orders": result["total_orders"],
        "delivered_orders": result["delivered_orders"],
        "delivery_rate": f"{result['delivered_orders'] / result['total_orders'] * 100:.1f}%",
        "total_income": round(result["total_income"], 2),
        "total_fixed_cost": round(result["total_fixed_cost"], 2),
        "total_swap_cost": round(result["total_swap_cost"], 2),
        "total_swaps": result["total_swaps"],
        "unfulfilled_penalty": round(result["unfulfilled_penalty"], 2),
        "total_cost": round(result["total_cost"], 2),
        "net_profit": round(result["net_profit"], 2),
        "total_distance_km": round(result["total_distance_km"], 2),
        "plan_file": plan_file,
        "detail_file": detail_file,
        "drone_stats": {},
    }
    for drone_id, ds in result["drone_stats"].items():
        summary["drone_stats"][drone_id] = {
            "total_deliveries": ds["total_deliveries"],
            "total_income": round(ds["total_income"], 2),
            "total_distance_km": round(ds["total_distance_km"], 2),
            "total_swaps": ds["total_swaps"],
            "final_location": ds["final_location"],
            "final_battery": round(ds["final_battery"], 1),
        }

    summary_file = os.path.join(output_dir, f"调度结果_{base_name}.json")
    with open(summary_file, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  结果摘要: {summary_file}")

    print("\n" + "=" * 70)
    print("  调度完成！")
    print("=" * 70)

    return result


def export_drone_flight_plans(result: Dict, output_dir: str, base_name: str):
    """导出3个独立的无人机飞行计划Excel文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "时间(分:秒)",
        "起降场",
        "动作",
        "参数",
        "装货订单",
        "卸货订单",
        "是否换电",
        "备注",
    ]

    for drone_id in ["UAV-01", "UAV-02", "UAV-03"]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = drone_id

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        actions = result["actions"].get(drone_id, [])
        for i, action in enumerate(actions, 2):

            def ga(a, key, default=""):
                if isinstance(a, dict):
                    return a.get(key, default)
                return getattr(a, key, default)

            action_type = ga(action, "action_type")
            time_str = seconds_to_time_str(ga(action, "time_seconds", 0))
            location = ga(action, "location", "")
            load_str = ",".join(str(oid) for oid in (ga(action, "load_orders") or []))
            unload_str = ",".join(
                str(oid) for oid in (ga(action, "unload_orders") or [])
            )
            swap_str = "是" if ga(action, "swap_battery", False) else "否"
            remark = ga(action, "remark", "")

            target_cruise = ga(action, "target_cruise_coord")
            parameters = ""
            if target_cruise:
                try:
                    c = target_cruise
                    parameters = f"{c.lng:.6f},{c.lat:.6f},{int(c.alt)}"
                except:
                    pass

            battery_before = ga(action, "battery_before", 0)
            battery_after = ga(action, "battery_after", 0)
            action_name = "起飞" if action_type == "takeoff" else "降落"
            if action_type == "landing":
                if ga(action, "swap_battery", False):
                    remark += f" 换电(电量{battery_before:.0f}%->100%)"
                else:
                    remark += f" 电量{battery_after:.0f}%"

            row_data = [
                time_str,
                location,
                action_name,
                parameters,
                load_str if action_type == "takeoff" else "",
                unload_str if action_type == "landing" else "",
                swap_str,
                remark.strip(),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center_align

        col_widths = [12, 16, 8, 30, 15, 15, 10, 25]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        filepath = os.path.join(output_dir, f"{drone_id}_飞行计划_{base_name}.xlsx")
        wb.save(filepath)
        print(f"  {drone_id}飞行计划: {filepath}")


def export_flight_plan(result: Dict, filepath: str):
    """导出飞行计划到Excel（按时间排序的合并计划）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "飞行计划"

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "时间(分:秒)",
        "起降场",
        "动作",
        "参数",
        "装货订单",
        "卸货订单",
        "是否换电",
        "备注",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = 2
    all_actions = []
    for drone_id in ["UAV-01", "UAV-02", "UAV-03"]:
        all_actions.extend((a, drone_id) for a in result["actions"].get(drone_id, []))
    
    all_actions.sort(key=lambda x: x[0].get("time_seconds", 0) if isinstance(x[0], dict) else x[0].time_seconds)

    for action, drone_id in all_actions:
        def get_attr(a, key, default=""):
            if isinstance(a, dict):
                return a.get(key, default)
            return getattr(a, key, default)

        action_type = get_attr(action, "action_type")
        time_seconds = get_attr(action, "time_seconds", 0)
        location = get_attr(action, "location", "")
        load_orders = get_attr(action, "load_orders", [])
        unload_orders = get_attr(action, "unload_orders", [])
        swap_battery = get_attr(action, "swap_battery", False)
        target_cruise = get_attr(action, "target_cruise_coord")
        battery_before = get_attr(action, "battery_before", 0)
        battery_after = get_attr(action, "battery_after", 0)
        remark = get_attr(action, "remark", "")

        time_str = seconds_to_time_str(time_seconds)
        load_str = ",".join(str(oid) for oid in (load_orders or []))
        unload_str = ",".join(str(oid) for oid in (unload_orders or []))
        swap_str = "是" if swap_battery else "否"
        parameters = ""
        if target_cruise:
            try:
                c = target_cruise
                parameters = f"{c.lng:.6f},{c.lat:.6f},{int(c.alt)}"
            except:
                pass

        action_name = "起飞" if action_type == "takeoff" else "降落"
        if action_type == "landing":
            if swap_battery:
                remark += f" 换电(电量{battery_before:.0f}%->100%)"
            else:
                remark += f" 电量{battery_after:.0f}%"

        row_data = [
            time_str,
            location,
            action_name,
            parameters,
            load_str if action_type == "takeoff" else "",
            unload_str if action_type == "landing" else "",
            swap_str,
            f"[{drone_id}] {remark.strip()}",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
        row_idx += 1

    col_widths = [12, 16, 8, 30, 15, 15, 10, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(filepath)


def export_assigned_orders(result: Dict, order_manager, filepath: str):
    """导出已分配订单列表到文本文件（按订单号从小到大排序）"""
    assigned_order_ids = set()
    for drone_actions in result.get("actions", {}).values():
        for action in drone_actions:
            load_orders = action.get("load_orders", []) if isinstance(action, dict) else getattr(action, "load_orders", [])
            for oid in load_orders:
                assigned_order_ids.add(oid)
    
    unassigned = set(order_manager.orders.keys()) - assigned_order_ids
    
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("=" * 50 + "\n")
        f.write("已分配订单列表（按订单号从小到大排序）\n")
        f.write("=" * 50 + "\n\n")
        
        f.write(f"总已分配订单数: {len(assigned_order_ids)}\n")
        f.write(f"总未分配订单数: {len(unassigned)}\n\n")
        
        f.write("-" * 50 + "\n")
        f.write("已分配订单（从小到大）:\n")
        f.write("-" * 50 + "\n")
        for oid in sorted(assigned_order_ids):
            order = order_manager.orders.get(oid)
            if order:
                f.write(f"  {oid}: {order.order_type.value} {order.weight_kg}kg\n")
            else:
                f.write(f"  {oid}\n")
        
        f.write("\n")
        f.write("-" * 50 + "\n")
        f.write("未分配订单（从小到大）:\n")
        f.write("-" * 50 + "\n")
        for oid in sorted(unassigned):
            order = order_manager.orders.get(oid)
            if order:
                f.write(f"  {oid}: {order.order_type.value} {order.weight_kg}kg\n")
            else:
                f.write(f"  {oid}\n")


def export_details(result: Dict, filepath: str, order_manager=None):
    """导出详细明细到Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()

    # 样式
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    sub_header_fill = PatternFill(
        start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
    )
    sub_header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    money_format = "#,##0.00"

    def write_header(ws, row, headers, fill=header_fill, font=header_font_white):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    # ============================================================
    # Sheet 1: 总体盈利摘要
    # ============================================================
    ws = wb.active
    ws.title = "盈利摘要"

    # 总体数据
    data_rows = [
        ("指标", "数值", "说明"),
        ("总订单数", result["total_orders"], "赛局中所有订单"),
        ("已配送订单数", result["delivered_orders"], "成功送达的订单"),
        (
            "配送率",
            f"{result['delivered_orders'] / result['total_orders'] * 100:.1f}%",
            "已配送/总订单",
        ),
        ("", "", ""),
        ("收入明细", "", ""),
        (
            "  订单配送收入",
            result["total_income"],
            "已送达订单的实际收入（扣迟到惩罚后）",
        ),
        ("", "", ""),
        ("成本明细", "", ""),
        (
            "  无人机固定成本",
            -result["total_fixed_cost"],
            f"{DRONE_COUNT}架 × {DRONE_FIXED_COST}元/架",
        ),
        (
            "  换电成本",
            -result["total_swap_cost"],
            f"{result['total_swaps']}次 × {BATTERY_SWAP_COST}元/次",
        ),
        ("  未履约惩罚", -result["unfulfilled_penalty"], "已抢但未送达的订单惩罚"),
        ("", "", ""),
        ("总计", "", ""),
        ("  总成本", -result["total_cost"], "固定成本+换电成本+惩罚"),
        ("  净利润", result["net_profit"], "总收入-总成本"),
        ("", "", ""),
        ("运营数据", "", ""),
        ("  总飞行距离(km)", f"{result['total_distance_km']:.2f}", "3架无人机累计"),
        ("  总换电次数", result["total_swaps"], "3架无人机累计"),
    ]

    for i, (label, value, desc) in enumerate(data_rows, 1):
        ws.cell(row=i, column=1, value=label).border = thin_border
        c2 = ws.cell(row=i, column=2, value=value)
        c2.border = thin_border
        c2.alignment = center_align
        ws.cell(row=i, column=3, value=desc).border = thin_border

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    # ============================================================
    # Sheet 2: 各无人机明细
    # ============================================================
    ws2 = wb.create_sheet("各无人机明细")
    headers2 = [
        "无人机",
        "最终位置",
        "配送订单数",
        "配送收入(元)",
        "固定成本(元)",
        "换电次数",
        "换电成本(元)",
        "飞行距离(km)",
        "最终电量(%)",
        "净利润(元)",
    ]
    write_header(ws2, 1, headers2)

    for i, (drone_id, ds) in enumerate(result["drone_stats"].items(), 2):
        net = ds["total_income"] - DRONE_FIXED_COST - ds["swap_cost"]
        row = [
            drone_id,
            ds["final_location"],
            ds["total_deliveries"],
            ds["total_income"],
            -DRONE_FIXED_COST,
            ds["total_swaps"],
            -ds["swap_cost"],
            f"{ds['total_distance_km']:.2f}",
            f"{ds['final_battery']:.1f}",
            net,
        ]
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    for col in range(1, 11):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # ============================================================
    # Sheet 3: 订单配送明细
    # ============================================================
    ws3 = wb.create_sheet("订单配送明细")
    headers3 = [
        "订单编号",
        "订单类型",
        "供应地",
        "需求地",
        "重量(kg)",
        "完整收入(元)",
        "实际收入(元)",
        "迟到惩罚(元)",
        "送达时刻(min)",
        "截止时刻(min)",
        "迟到(min)",
    ]
    write_header(ws3, 1, headers3)

    # 按订单编号排序
    sorted_details = sorted(result["income_details"], key=lambda x: x["order_id"])

    for i, detail in enumerate(sorted_details, 2):
        late_penalty = detail["full_income"] - detail["actual_income"]
        row = [
            detail["order_id"],
            detail["order_type"],
            detail["supply"],
            detail["demand"],
            detail["weight_kg"],
            detail["full_income"],
            detail["actual_income"],
            -late_penalty,
            f"{detail['delivery_time_min']:.1f}",
            f"{detail['deadline_min']:.1f}",
            f"{detail['late_minutes']:.1f}",
        ]
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    for col in range(1, 12):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # ============================================================
    # Sheet 4: 按订单类型统计
    # ============================================================
    ws4 = wb.create_sheet("按类型统计")
    headers4 = [
        "订单类型",
        "配送数量",
        "完整收入(元)",
        "实际收入(元)",
        "迟到损失(元)",
        "迟到单数",
        "收入实现率(%)",
    ]
    write_header(ws4, 1, headers4)

    for i, (otype, ts) in enumerate(sorted(result["type_stats"].items()), 2):
        loss = ts["full_income"] - ts["actual_income"]
        rate = (
            ts["actual_income"] / ts["full_income"] * 100
            if ts["full_income"] > 0
            else 0
        )
        row = [
            otype,
            ts["count"],
            ts["full_income"],
            ts["actual_income"],
            -loss,
            ts["late_count"],
            f"{rate:.1f}",
        ]
        for col, val in enumerate(row, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    for col in range(1, 8):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # ============================================================
    # Sheet 5: 按路线统计
    # ============================================================
    ws5 = wb.create_sheet("按路线统计")
    headers5 = ["路线", "配送数量", "实际收入(元)", "平均收入(元/单)"]
    write_header(ws5, 1, headers5)

    sorted_routes = sorted(
        result["route_stats"].items(), key=lambda x: -x[1]["actual_income"]
    )
    for i, (route, rs) in enumerate(sorted_routes, 2):
        avg = rs["actual_income"] / rs["count"] if rs["count"] > 0 else 0
        row = [route, rs["count"], rs["actual_income"], f"{avg:.2f}"]
        for col, val in enumerate(row, 1):
            cell = ws5.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    for col in range(1, 5):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # ============================================================
    # Sheet 5.5: 订单分配明细
    # ============================================================
    ws5b = wb.create_sheet("订单分配明细")
    headers5b = ["订单号", "类型", "重量(kg)", "供应地", "需求地", "是否已分配", "送达时间"]
    write_header(ws5b, 1, headers5b)
    
    assigned_order_ids = set()
    delivery_times = {}
    for drone_actions in result.get("actions", {}).values():
        for action in drone_actions:
            load_orders = action.get("load_orders", []) if isinstance(action, dict) else getattr(action, "load_orders", [])
            for oid in load_orders:
                assigned_order_ids.add(oid)
            if action.get("action_type") == "landing" or (isinstance(action, dict) and action.get("action_type") == "landing"):
                unload_orders = action.get("unload_orders", [])
                time_seconds = action.get("time_seconds", 0)
                for oid in unload_orders:
                    if oid not in delivery_times:
                        delivery_times[oid] = time_seconds
    
    unassigned_ids = set(order_manager.orders.keys()) - assigned_order_ids
    
    row = 2
    for oid in sorted(order_manager.orders.keys()):
        order = order_manager.orders[oid]
        is_assigned = oid in assigned_order_ids
        delivery_time = delivery_times.get(oid)
        delivery_str = f"{int(delivery_time // 60)}:{int(delivery_time % 60):02d}" if delivery_time is not None else "未分配"
        
        row_data = [
            oid,
            order.order_type.value if hasattr(order.order_type, 'value') else str(order.order_type),
            order.weight_kg,
            order.supply_location,
            order.demand_location,
            "是" if is_assigned else "否",
            delivery_str,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws5b.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
        row += 1
    
    for col in range(1, 8):
        ws5b.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # ============================================================
    # Sheet 6: 各无人机飞行日志
    # ============================================================
    for drone_id in ["UAV-01", "UAV-02", "UAV-03"]:
        ws_drone = wb.create_sheet(f"{drone_id}日志")
        headers6 = [
            "时间",
            "动作",
            "位置",
            "装货订单",
            "卸货订单",
            "换电",
            "飞行距离(km)",
            "电量消耗(%)",
            "电量前(%)",
            "电量后(%)",
            "载重前(kg)",
            "载重后(kg)",
            "备注",
        ]
        write_header(ws_drone, 1, headers6)

        actions = result["actions"].get(drone_id, [])
        for i, action in enumerate(actions, 2):
            # Support both dict and FlightAction
            def ga(a, key, default=""):
                if isinstance(a, dict):
                    return a.get(key, default)
                return getattr(a, key, default)

            time_str = seconds_to_time_str(ga(action, "time_seconds", 0))
            load_str = ",".join(str(oid) for oid in (ga(action, "load_orders") or []))
            unload_str = ",".join(
                str(oid) for oid in (ga(action, "unload_orders") or [])
            )
            swap_str = "是" if ga(action, "swap_battery", False) else "否"
            flight_dist = ga(action, "flight_distance_km", 0)
            bat_before = ga(action, "battery_before", 0)
            bat_after = ga(action, "battery_after", 0)
            payload_before = ga(action, "payload_before", 0)
            payload_after = ga(action, "payload_after", 0)
            remark = ga(action, "remark", "")
            bat_consumed = bat_before - bat_after if bat_before and bat_after else 0

            row = [
                time_str,
                ga(action, "action_type", ""),
                ga(action, "location", ""),
                load_str,
                unload_str,
                swap_str,
                f"{flight_dist:.3f}" if flight_dist > 0 else "",
                f"{bat_consumed:.1f}" if bat_consumed > 0 else "",
                f"{bat_before:.1f}" if bat_before else "",
                f"{bat_after:.1f}" if bat_after else "",
                f"{payload_before:.1f}" if payload_before else "",
                f"{payload_after:.1f}" if payload_after else "",
                remark,
            ]
            for col, val in enumerate(row, 1):
                cell = ws_drone.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center_align

        for col in range(1, 14):
            ws_drone.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    wb.save(filepath)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="无人机配送调度系统（ALNS版）")
    parser.add_argument(
        "--input",
        "-i",
        type=str,
        default=None,
        help="订单数据Excel文件路径（默认: input/input_data.xlsx）",
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None, help="输出目录（默认: output/）"
    )
    parser.add_argument(
        "--iterations",
        "-n",
        type=int,
        default=5000,
        help="ALNS最大迭代次数（默认: 5000）",
    )
    parser.add_argument(
        "--time-limit",
        "-t",
        type=float,
        default=120.0,
        help="ALNS时间限制（秒，默认: 120）",
    )
    parser.add_argument(
        "--max-orders", type=int, default=600, help="初始解最大分配订单数（默认: 600）"
    )
    args = parser.parse_args()

    try:
        run(
            data_file=args.input,
            output_dir=args.output,
            iterations=args.iterations,
            time_limit=args.time_limit,
            max_orders=args.max_orders,
        )
    except Exception as e:
        error_info = {
            "status": "error",
            "message": str(e),
        }
        output_dir = args.output if args.output else os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "output"
        )
        error_file = os.path.join(output_dir, "调度结果_error.json")
        os.makedirs(output_dir, exist_ok=True)
        with open(error_file, "w", encoding="utf-8") as f:
            json.dump(error_info, f, ensure_ascii=False, indent=2)
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        sys.exit(1)
