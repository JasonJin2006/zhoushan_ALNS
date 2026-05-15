"""
订单处理模块
============
5种订单类型（海鲜A/B/C、日用品D/E），收入40元/kg，
各有不同的重量、时间窗和迟到惩罚率。
"""

from dataclasses import dataclass, field
from enum import Enum
from typing import Dict, List, Optional, Tuple
import openpyxl


# ============================================================
# 订单类型定义
# ============================================================

class OrderType(Enum):
    """订单类型枚举"""
    SEAFOOD_A = "海鲜A"   # 从渔船运往陆地和轮船
    SEAFOOD_B = "海鲜B"   # 从渔船运往轮船
    SEAFOOD_C = "海鲜C"   # 从渔船运往轮船（高价值高时效）
    DAILY_D = "日用品D"    # 从陆地运往渔船和轮船
    DAILY_E = "日用品E"    # 从轮船运往渔船


# 订单类型属性映射
ORDER_TYPE_PROPERTIES = {
    # type: (中文名, 重量kg, 时间窗min, 迟到惩罚元/min, 描述)
    OrderType.SEAFOOD_A: ("海鲜A", 1.0, 40, 2, "从渔船运往陆地和轮船"),
    OrderType.SEAFOOD_B: ("海鲜B", 1.5, 30, 4, "从渔船运往轮船"),
    OrderType.SEAFOOD_C: ("海鲜C", 2.5, 15, 5, "从渔船运往轮船（高时效）"),
    OrderType.DAILY_D:   ("日用品D", 1.5, 60, 4, "从陆地运往渔船和轮船"),
    OrderType.DAILY_E:   ("日用品E", 1.0, 40, 1, "从轮船运往渔船"),
}

# 收入标准：40元/kg
INCOME_PER_KG = 40.0

# 数字类型到枚举的映射（用于读取Excel数据）
NUMERIC_TYPE_MAP = {
    '1': OrderType.SEAFOOD_A,
    '2': OrderType.SEAFOOD_B,
    '3': OrderType.SEAFOOD_C,
    '4': OrderType.DAILY_D,
    '5': OrderType.DAILY_E,
}

# 中文名到枚举的映射
CHINESE_TYPE_MAP = {
    '海鲜A': OrderType.SEAFOOD_A,
    '海鲜B': OrderType.SEAFOOD_B,
    '海鲜C': OrderType.SEAFOOD_C,
    '日用品D': OrderType.DAILY_D,
    '日用品E': OrderType.DAILY_E,
}


# ============================================================
# 订单数据结构
# ============================================================

@dataclass
class Order:
    """订单"""
    order_id: int               # 订单编号
    time_sequence: int          # 时间序号（批次）
    generation_time_seconds: float  # 生成时间（秒）
    generation_time_str: str    # 生成时间原始字符串
    supply_location: str        # 供应地
    demand_location: str        # 需求地
    order_type: OrderType       # 订单类型
    weight_kg: float            # 重量（kg）
    time_window_minutes: float  # 时间窗（分钟）
    late_penalty_rate: float    # 迟到惩罚率（元/分钟）
    quantity: int = 1           # 数量（通常为1）

    @property
    def full_income(self) -> float:
        """完整收入（时间窗内送达）= 40元/kg × 重量"""
        return INCOME_PER_KG * self.weight_kg

    @property
    def deadline_seconds(self) -> float:
        """时间窗截止时刻（秒）= 生成时间 + 时间窗"""
        return self.generation_time_seconds + self.time_window_minutes * 60

    @property
    def income_zero_seconds(self) -> float:
        """收入降为0的时刻（秒）"""
        # 收入降为0时：full_income - late_minutes * penalty_rate = 0
        # late_minutes = full_income / penalty_rate
        if self.late_penalty_rate > 0:
            late_minutes_to_zero = self.full_income / self.late_penalty_rate
            return self.deadline_seconds + late_minutes_to_zero * 60
        return float('inf')

    @property
    def unfulfilled_seconds(self) -> float:
        """订单变为未完成的时刻（秒）
        超过此时刻仍未送达则视为未完成。
        这与income_zero_seconds相同——收入降为0后即视为未完成。"""
        return self.income_zero_seconds

    def income_at_delivery(self, delivery_time_seconds: float) -> float:
        """
        计算在指定时刻送达时的收入。

        - 在时间窗内送达: 完整收入
        - 迟到但收入未降为0: 收入 - 迟到惩罚
        - 超过未完成时刻: 未完成，收入为0，需额外惩罚
        """
        if delivery_time_seconds <= self.deadline_seconds:
            # 时间窗内送达
            return self.full_income
        elif delivery_time_seconds <= self.income_zero_seconds:
            # 迟到但还未变成未完成
            late_minutes = (delivery_time_seconds - self.deadline_seconds) / 60.0
            penalty = late_minutes * self.late_penalty_rate
            return max(0, self.full_income - penalty)
        else:
            # 未完成
            return 0.0

    def penalty_if_unfulfilled(self) -> float:
        """
        如果订单未完成（未送达），的惩罚金额。
        未履约惩罚: -40元/kg × 重量
        """
        return UNFULFILLED_PENALTY_PER_KG * self.weight_kg

    def is_deliverable_at(self, current_time_seconds: float) -> bool:
        """
        判断在当前时刻是否还能送达（收入是否大于0）。
        如果已经过了收入降为0的时刻，则不值得再送。
        """
        return current_time_seconds < self.income_zero_seconds

    def urgency_score(self, current_time_seconds: float) -> float:
        """
        计算订单的紧急程度得分。
        得分越高越紧急：距离截止时间越近、惩罚率越高、重量越大越紧急。
        """
        remaining = max(0, self.deadline_seconds - current_time_seconds)
        if remaining == 0:
            # 已超时，按收入递减速率衡量
            remaining_to_zero = max(0, self.income_zero_seconds - current_time_seconds)
            if remaining_to_zero == 0:
                return 0  # 已无法获得收入
            return 1000 + self.late_penalty_rate * self.weight_kg / (remaining_to_zero / 60 + 1)
        return (self.late_penalty_rate * self.weight_kg) / (remaining / 60 + 1)


# ============================================================
# 订单管理器
# ============================================================

class OrderManager:
    """
    订单管理器
    ==========
    管理所有订单的加载、查询和状态跟踪。
    """

    # 未履约惩罚
    UNFULFILLED_PENALTY_PER_KG = -40.0

    def __init__(self):
        self.orders: Dict[int, Order] = {}  # order_id -> Order
        self.grabbed_orders: Dict[int, Order] = {}  # 已抢单
        self.delivered_orders: Dict[int, Tuple[Order, float]] = {}  # 已送达: (order, delivery_time)
        self.unfulfilled_orders: Dict[int, Order] = {}  # 未完成
        self.max_pending_orders = 90  # 最大待配送订单数

    def load_from_excel(self, filepath: str, sheet_name: str = '完整版',
                        type_column: int = 5) -> int:
        """
        从Excel文件加载订单数据。

        参数:
            filepath: Excel文件路径
            sheet_name: 工作表名称
            type_column: 订单种类所在列索引（0-based）

        返回:
            加载的订单数量
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # 尝试指定的工作表名称，如果不存在则尝试常见替代名称
        available_sheets = wb.sheetnames
        target_sheet = sheet_name
        if target_sheet not in available_sheets:
            # 尝试常见的替代名称
            alternatives = ['完整版', '订单任务', '订单']
            for alt in alternatives:
                if alt in available_sheets:
                    target_sheet = alt
                    break
            else:
                # 使用最后一个sheet（通常是订单数据）
                target_sheet = available_sheets[-1]
        ws = wb[target_sheet]

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            try:
                order = self._parse_order_row(row, type_column)
                self.orders[order.order_id] = order
                count += 1
            except (ValueError, KeyError) as e:
                print(f"Warning: 跳过无效订单行: {row}, 错误: {e}")
                continue

        wb.close()
        return count

    def _parse_order_row(self, row: tuple, type_column: int) -> Order:
        """解析一行订单数据"""
        time_seq = int(row[0])
        order_id = int(row[1])
        gen_time_str = str(row[2])
        supply = str(row[3])
        demand = str(row[4])
        order_type_raw = str(row[5])
        weight = float(row[6])
        time_window = float(row[7])
        late_penalty = float(row[8])
        quantity = int(row[9]) if row[9] is not None else 1

        # 解析订单类型
        order_type = CHINESE_TYPE_MAP.get(order_type_raw) or NUMERIC_TYPE_MAP.get(order_type_raw)
        if order_type is None:
            raise ValueError(f"未知订单类型: {order_type_raw}")

        # 解析生成时间（格式如 "0:12", "178:08"）
        gen_seconds = self._parse_time_str(gen_time_str)

        return Order(
            order_id=order_id,
            time_sequence=time_seq,
            generation_time_seconds=gen_seconds,
            generation_time_str=gen_time_str,
            supply_location=supply,
            demand_location=demand,
            order_type=order_type,
            weight_kg=weight,
            time_window_minutes=time_window,
            late_penalty_rate=late_penalty,
            quantity=quantity,
        )

    @staticmethod
    def _parse_time_str(time_str: str) -> float:
        """
        解析时间字符串为秒数。
        格式: "分:秒" 如 "0:12" = 12秒, "178:08" = 178分8秒 = 10688秒
        """
        parts = time_str.strip().split(':')
        minutes = int(parts[0])
        seconds = int(parts[1]) if len(parts) > 1 else 0
        return minutes * 60 + seconds

    def grab_order(self, order_id: int) -> bool:
        """
        抢单操作。将订单从可用池移到已抢池。

        返回:
            是否成功抢单
        """
        if order_id not in self.orders:
            return False
        if order_id in self.grabbed_orders:
            return False
        # 检查待配送订单数限制
        pending = len(self.grabbed_orders) - len(self.delivered_orders) - len(self.unfulfilled_orders)
        if pending >= self.max_pending_orders:
            return False
        self.grabbed_orders[order_id] = self.orders[order_id]
        return True

    def deliver_order(self, order_id: int, delivery_time_seconds: float) -> float:
        """
        完成订单配送。

        参数:
            order_id: 订单编号
            delivery_time_seconds: 送达时刻（秒）

        返回:
            订单收入（可能为0或负值）
        """
        if order_id not in self.grabbed_orders:
            return 0.0

        order = self.grabbed_orders[order_id]
        income = order.income_at_delivery(delivery_time_seconds)

        if income > 0:
            self.delivered_orders[order_id] = (order, delivery_time_seconds)
        else:
            # 超时未完成
            self.unfulfilled_orders[order_id] = order

        del self.grabbed_orders[order_id]
        return income

    def mark_unfulfilled(self, order_id: int) -> float:
        """
        标记订单为未完成。

        返回:
            惩罚金额（负数）
        """
        if order_id not in self.grabbed_orders:
            return 0.0

        order = self.grabbed_orders[order_id]
        penalty = order.penalty_if_unfulfilled()
        self.unfulfilled_orders[order_id] = order
        del self.grabbed_orders[order_id]
        return penalty

    # ---- 查询接口 ----

    def get_orders_by_supply(self, supply_location: str) -> List[Order]:
        """获取指定供应地的所有订单"""
        return [o for o in self.orders.values() if o.supply_location == supply_location]

    def get_orders_by_demand(self, demand_location: str) -> List[Order]:
        """获取指定需求地的所有订单"""
        return [o for o in self.orders.values() if o.demand_location == demand_location]

    def get_orders_by_type(self, order_type: OrderType) -> List[Order]:
        """获取指定类型的所有订单"""
        return [o for o in self.orders.values() if o.order_type == order_type]

    def get_orders_at_time(self, time_seconds: float) -> List[Order]:
        """获取在指定时刻已生成但未超时的订单"""
        return [
            o for o in self.orders.values()
            if o.generation_time_seconds <= time_seconds and o.is_deliverable_at(time_seconds)
        ]

    def get_pending_orders(self) -> List[Order]:
        """获取当前待配送的已抢订单"""
        return list(self.grabbed_orders.values())

    def get_urgent_orders(self, current_time_seconds: float, limit: int = 20) -> List[Order]:
        """获取最紧急的待配送订单"""
        pending = self.get_pending_orders()
        scored = [(o.urgency_score(current_time_seconds), o) for o in pending]
        scored.sort(key=lambda x: -x[0])
        return [o for _, o in scored[:limit]]

    def get_supply_demand_pairs(self) -> Dict[str, Dict[str, int]]:
        """获取供应地->需求地的订单数量分布"""
        result: Dict[str, Dict[str, int]] = {}
        for o in self.orders.values():
            if o.supply_location not in result:
                result[o.supply_location] = {}
            if o.demand_location not in result[o.supply_location]:
                result[o.supply_location][o.demand_location] = 0
            result[o.supply_location][o.demand_location] += 1
        return result

    def get_type_statistics(self) -> Dict[OrderType, int]:
        """获取各类型订单数量统计"""
        stats: Dict[OrderType, int] = {}
        for o in self.orders.values():
            stats[o.order_type] = stats.get(o.order_type, 0) + 1
        return stats

    def get_total_potential_revenue(self) -> float:
        """计算所有订单如果全部按时送达的总收入"""
        return sum(o.full_income for o in self.orders.values())

    # ---- 收入统计 ----

    def get_delivered_revenue(self) -> float:
        """获取已送达订单的总收入"""
        total = 0.0
        for order, delivery_time in self.delivered_orders.values():
            total += order.income_at_delivery(delivery_time)
        return total

    def get_unfulfilled_penalty(self) -> float:
        """获取未完成订单的总惩罚"""
        return sum(o.penalty_if_unfulfilled() for o in self.unfulfilled_orders.values())

    def get_current_pending_penalty_if_all_fail(self) -> float:
        """如果当前所有待配送订单都未完成的总惩罚"""
        return sum(o.penalty_if_unfulfilled() for o in self.grabbed_orders.values())
